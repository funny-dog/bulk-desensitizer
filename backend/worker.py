from __future__ import annotations

import csv
import json
from pathlib import Path

import redis
from celery import Task
from openpyxl import Workbook, load_workbook

from celery_app import celery_app
from config import settings
from database import SessionLocal, init_db
from models import DataRecord

# Initialize Redis client for Pub/Sub
redis_client = redis.from_url(settings.celery_broker_url)

EMAIL_KEYWORDS = {"email", "e-mail", "mail", "邮箱"}
PHONE_KEYWORDS = {"phone", "mobile", "tel", "telephone", "手机号", "电话"}
ID_KEYWORDS = {"id_card", "idcard", "identity", "ssn", "passport", "身份证", "证件"}
NAME_KEYWORDS = {"name", "full_name", "first_name", "last_name", "姓名"}
ADDRESS_KEYWORDS = {"address", "addr", "地址"}


def _iter_csv_rows(path: Path):
    with path.open("r", newline="") as handle:
        reader = csv.reader(handle)
        yield from reader


def _iter_xlsx_rows(path: Path):
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            yield row
    finally:
        workbook.close()


def _count_rows(path: Path, file_type: str) -> int:
    if file_type == "csv":
        return sum(1 for _ in _iter_csv_rows(path))
    if file_type == "xlsx":
        return sum(1 for _ in _iter_xlsx_rows(path))
    raise ValueError(f"unsupported file type: {file_type}")


def _iter_rows(path: Path, file_type: str):
    if file_type == "csv":
        return _iter_csv_rows(path)
    if file_type == "xlsx":
        return _iter_xlsx_rows(path)
    raise ValueError(f"unsupported file type: {file_type}")


def _row_payload(row) -> str:
    return ",".join("" if cell is None else str(cell) for cell in row)


def _select_masker(header: str):
    normalized = header.strip().lower()
    if any(key in normalized for key in EMAIL_KEYWORDS):
        return _mask_email
    if any(key in normalized for key in PHONE_KEYWORDS):
        return _mask_phone
    if any(key in normalized for key in ID_KEYWORDS):
        return _mask_id
    if any(key in normalized for key in NAME_KEYWORDS):
        return _mask_name
    if any(key in normalized for key in ADDRESS_KEYWORDS):
        return _mask_address
    return None


def _mask_email(value: str) -> str:
    if "@" not in value:
        return _mask_generic(value)
    local, domain = value.split("@", 1)
    if not local:
        return f"***@{domain}"
    return f"{local[0]}***@{domain}"


def _mask_phone(value: str) -> str:
    digits = [char for char in value if char.isdigit()]
    if not digits:
        return _mask_generic(value)
    keep = 4 if len(digits) > 4 else 0
    masked = []
    seen = 0
    for char in value:
        if char.isdigit():
            seen += 1
            if len(digits) - seen < keep:
                masked.append(char)
            else:
                masked.append("*")
        else:
            masked.append(char)
    return "".join(masked)


def _mask_id(value: str) -> str:
    if len(value) <= 4:
        return "*" * len(value)
    return f"{value[:2]}{'*' * (len(value) - 4)}{value[-2:]}"


def _mask_name(value: str) -> str:
    if len(value) <= 1:
        return "*" * len(value)
    return f"{value[0]}{'*' * (len(value) - 1)}"


def _mask_address(value: str) -> str:
    if len(value) <= 6:
        return "*" * len(value)
    return f"{value[:6]}{'*' * (len(value) - 6)}"


def _mask_generic(value: str) -> str:
    if len(value) <= 2:
        return "*" * len(value)
    return f"{value[0]}{'*' * (len(value) - 2)}{value[-1]}"


def _apply_mask(value, masker):
    if value is None:
        return ""
    text = str(value)
    if not text:
        return text
    return masker(text) if masker else text


@celery_app.task(bind=True)
def process_csv(self: Task, file_path: str) -> dict:
    init_db()
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"file not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError(f"unsupported file type: {suffix}")
    file_type = "xlsx" if suffix == ".xlsx" else "csv"

    total = _count_rows(path, file_type)

    if total == 0:
        return {"current": 0, "total": 0, "message": "no rows"}

    session = SessionLocal()
    try:
        for index, row in enumerate(_iter_rows(path, file_type), start=1):
            record = DataRecord(
                task_id=self.request.id,
                row_number=index,
                payload=_row_payload(row),
            )
            session.add(record)

            if index % 100 == 0:
                session.commit()

            if index == total or index % 10 == 0:
                self.update_state(
                    state="PROGRESS",
                    meta={
                        "current": index,
                        "total": total,
                        "message": f"Processing row {index}/{total}",
                    },
                )
                # Publish progress to Redis
                redis_client.publish(
                    f"task_progress:{self.request.id}",
                    json.dumps(
                        {
                            "current": index,
                            "total": total,
                            "message": f"Processing row {index}/{total}",
                        }
                    ),
                )

        session.commit()
    finally:
        session.close()

    # Publish completion message
    redis_client.publish(
        f"task_progress:{self.request.id}",
        json.dumps({"current": total, "total": total, "message": "completed"}),
    )

    return {"current": total, "total": total, "message": "completed"}


@celery_app.task(bind=True)
def process_desensitize(self: Task, file_path: str) -> dict:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"file not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError(f"unsupported file type: {suffix}")
    file_type = "xlsx" if suffix == ".xlsx" else "csv"

    output_dir = Path(settings.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{self.request.id}_desensitized{suffix}"

    total_rows = _count_rows(path, file_type)
    rows = _iter_rows(path, file_type)
    try:
        header = next(rows)
    except StopIteration:
        return {"current": 0, "total": 0, "message": "no rows"}

    header_cells = ["" if cell is None else str(cell) for cell in header]
    maskers = [_select_masker(cell) for cell in header_cells]

    data_total = max(total_rows - 1, 0)

    if file_type == "csv":
        with output_path.open("w", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(header_cells)
            for index, row in enumerate(rows, start=1):
                masked_row = []
                for i, cell in enumerate(row):
                    masker = maskers[i] if i < len(maskers) else None
                    masked_row.append(_apply_mask(cell, masker))
                writer.writerow(masked_row)

                if data_total and (index == data_total or index % 10 == 0):
                    message = f"Desensitizing row {index}/{data_total}"
                    self.update_state(
                        state="PROGRESS",
                        meta={"current": index, "total": data_total, "message": message},
                    )
                    redis_client.publish(
                        f"task_progress:{self.request.id}",
                        json.dumps({"current": index, "total": data_total, "message": message}),
                    )
    else:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet()
        sheet.append(header_cells)
        for index, row in enumerate(rows, start=1):
            masked_row = []
            for i, cell in enumerate(row):
                masker = maskers[i] if i < len(maskers) else None
                masked_row.append(_apply_mask(cell, masker))
            sheet.append(masked_row)

            if data_total and (index == data_total or index % 10 == 0):
                message = f"Desensitizing row {index}/{data_total}"
                self.update_state(
                    state="PROGRESS",
                    meta={"current": index, "total": data_total, "message": message},
                )
                redis_client.publish(
                    f"task_progress:{self.request.id}",
                    json.dumps({"current": index, "total": data_total, "message": message}),
                )
        workbook.save(output_path)
        workbook.close()

    if data_total == 0:
        redis_client.publish(
            f"task_progress:{self.request.id}",
            json.dumps({"current": 0, "total": 0, "message": "completed"}),
        )
        return {
            "current": 0,
            "total": 0,
            "message": "completed",
            "output_file": output_path.name,
        }

    redis_client.publish(
        f"task_progress:{self.request.id}",
        json.dumps({"current": data_total, "total": data_total, "message": "completed"}),
    )

    return {
        "current": data_total,
        "total": data_total,
        "message": "completed",
        "output_file": output_path.name,
    }
